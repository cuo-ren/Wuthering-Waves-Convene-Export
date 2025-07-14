import json
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import os
import time

def export_to_excel():
    os.makedirs("./export/excel", exist_ok=True)
    for uid in gacha_list:
        title_font = Font(name = "微软雅黑", bold=True, color = "757575")
        title_fill = PatternFill(fill_type="solid", fgColor="DBD7D3")
        content_fill = PatternFill(fill_type="solid", fgColor="EBEBEB")
        purple_font = Font(name = "微软雅黑", bold=True, color = "A256E1")
        gold_font = Font(name = "微软雅黑", bold=True, color = "BD6932")
        black_font = Font(name = "微软雅黑", color = "8E8E8E")
        thin_border = Border(
            left=Side(style="thin", color="C4C2BF"),
            right=Side(style="thin", color="C4C2BF"),
            top=Side(style="thin", color="C4C2BF"),
            bottom=Side(style="thin", color="C4C2BF")
        )
        wb = Workbook()
        wb.remove(wb.active)
        for key in gacha_list[uid]["data"]:
            for t in gacha_type["data"]:
                if t["key"] == key:
                    ws = wb.create_sheet(title=language[config["language"]][t["name"]])
                    break
            ws.append([language[config["language"]]["time"],language[config["language"]]["name"],language[config["language"]]["type"],language[config["language"]]["stars"],language[config["language"]]["TotalPulls"],language[config["language"]]["TotalPullsSince5"]])
            for cell in ws[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = thin_border
            total_count = 0
            count = 0
            for item in gacha_list[uid]["data"][key]:
                total_count += 1
                count += 1
                ws.append([item["time"],item["name"],item["type"],item["qualityLevel"],total_count,count])
                if item["qualityLevel"] == 4:
                    for cell in ws[total_count+1]:
                        cell.font = purple_font
                        cell.fill = content_fill
                        cell.border = thin_border
                elif item["qualityLevel"] == 5:
                    for cell in ws[total_count+1]:
                        cell.font = gold_font
                        cell.fill = content_fill
                        cell.border = thin_border
                    count = 0
                elif item["qualityLevel"] == 3:
                    for cell in ws[total_count+1]:
                        cell.font = black_font
                        cell.fill = content_fill
                        cell.border = thin_border
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                if col_letter not in ["A", "B", "F"]:
                    continue  # 跳过非目标列
                if col_letter == "A":
                    ws.column_dimensions[col_letter].width = 23.36
                if col_letter == "B":
                    ws.column_dimensions[col_letter].width = 15.73
                if col_letter == "F":
                    ws.column_dimensions[col_letter].width = 10.91

        wb.save(f"./export/excel/{language[config["language"]]["csvFilename"]}_{uid}_{int(time.time())}.xlsx")

with open("language.json",mode="r",encoding="utf-8") as f:
    language = json.load(f)

with open("config.json",mode="r",encoding="utf-8") as f:
    config = json.load(f)

with open("GachaType.json",mode="r",encoding="utf-8") as f:
     gacha_type = json.load(f)

with open("./data/gacha_list.json",mode="r",encoding="utf-8") as f:
     gacha_list = json.load(f)

export_to_excel()
